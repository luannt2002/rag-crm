"""ExcelOpenpyxlParser — REAL strategy for ``.xlsx`` ingestion.

Uses ``openpyxl`` (Apache 2.0, pure-Python) to walk every sheet and convert it
to ONE structured-markdown document (AdapChunk L1, via
``rows_to_structured_markdown``) — multi-table + section-title aware, so each
sub-table stays under its heading. A workbook with >1 tab nests each sheet under
a top-level ``# <sheet>`` heading. Same canonical form as the Sheets / Kreuzberg
parsers, so the downstream chunker/extractor is format-agnostic.

Domain-neutral: NO hardcoded sheet / column names — every label flows from
the workbook. ``openpyxl`` is an optional dep; init raises ``ImportError``
when missing so the registry's fail-soft path falls back to NullParser and
operators see the install hint in logs.
"""

from __future__ import annotations

from io import BytesIO

import structlog

from ragbot.shared.tabular_markdown import (
    rows_to_structured_markdown,
    split_markdown_to_row_chunks,
)

logger = structlog.get_logger(__name__)


_XLSX_MIME: str = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
_XLSX_EXT: str = ".xlsx"


def _openpyxl_available() -> bool:
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


class ExcelOpenpyxlParser:
    """Excel ``.xlsx`` parser — workbook → ONE structured-markdown document."""

    def __init__(self, **_: object) -> None:
        if not _openpyxl_available():
            raise ImportError(
                "openpyxl is not installed; add to pyproject or set "
                "system_config.document_parser_provider='null'."
            )

    @staticmethod
    def get_provider_name() -> str:
        return "excel_openpyxl"

    def supports(self, mime_type: str, file_ext: str) -> bool:
        return (
            (mime_type or "").strip().lower() == _XLSX_MIME
            or (file_ext or "").strip().lower() == _XLSX_EXT
        )

    async def parse(
        self,
        content: bytes,
        *,
        file_name: str,
    ) -> list[dict]:
        # Local import — module-level import would crash on systems without
        # openpyxl, defeating the registry's fail-soft fallback.
        from openpyxl import load_workbook

        wb = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        sheet_count = len(wb.sheetnames)
        parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [
                [("" if c is None else str(c)).strip() for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
            md = rows_to_structured_markdown(rows)
            if not md.strip():
                continue
            # Prefix each sheet as a top-level heading when the workbook has >1 tab
            # so sub-table sections nest under their sheet (a sheet IS a section too).
            if sheet_count > 1:
                parts.append(f"# {sheet_name}\n\n{md}")
            else:
                parts.append(md)

        wb.close()
        markdown = "\n\n".join(parts)
        if not markdown.strip():
            return []
        heading_lines = sum(
            1 for ln in markdown.splitlines() if ln.lstrip().startswith("#")
        )
        logger.info(
            "excel_openpyxl_parsed",
            file_name=file_name,
            sheets=sheet_count,
            markdown_chars=len(markdown),
            section_headings=heading_lines,
        )
        # ROW-AS-CHUNK parity with GoogleSheetsParser: split the section-bound
        # markdown into one atomic chunk per data row (header + section bound in),
        # so a large workbook is not embedded as one Lost-in-the-Middle blob and
        # the chunker never packs two rows together (cross-row value mis-bind).
        # Fallback to the whole doc if the split yields nothing (never drop
        # content). Stats extraction is unaffected — it re-reads the raw rows.
        row_chunks = split_markdown_to_row_chunks(markdown) or [markdown]
        base_meta = {
            "file_name": file_name,
            "parser": self.get_provider_name(),
            "format": "markdown",
            "section_headings": heading_lines,
        }
        return [{"content": c, "metadata": dict(base_meta)} for c in row_chunks]


__all__ = ["ExcelOpenpyxlParser"]
